import os
import math

from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QProgressBar, QLineEdit, QGroupBox,
    QAbstractItemView, QMessageBox, QCheckBox
)
from qgis.PyQt.QtCore import Qt, QThread, pyqtSignal
from qgis.PyQt.QtGui import QColor

from qgis.core import (
    QgsVectorLayer, QgsProject, QgsField,
    QgsFeature, QgsGeometry, QgsPointXY,
    QgsMarkerSymbol, QgsWkbTypes,
    QgsSingleSymbolRenderer
)
from qgis.PyQt.QtCore import QVariant


# ── Detection logic ───────────────────────────────────────────────

def force_2d_coords(coords):
    return [(c[0], c[1]) for c in coords]


def find_errors_in_layer(layer, angle_threshold=2.0,
                         check_spikes=True, check_nonconvex=True):
    results     = []
    field_names = [f.name() for f in layer.fields()]

    for feature in layer.getFeatures():
        geom = feature.geometry()
        if geom is None or geom.isEmpty():
            continue

        if QgsWkbTypes.isMultiType(geom.wkbType()):
            parts = geom.asMultiPolygon()
        else:
            parts = [geom.asPolygon()]

        fid  = feature.id()
        slno = feature["sl.no"]      if "sl.no"       in field_names else str(fid)
        code = feature["AiDashCode"] if "AiDashCode"  in field_names else "N/A"

        for part_idx, polygon in enumerate(parts):
            if not polygon:
                continue

            exterior = polygon[0]
            coords   = force_2d_coords(exterior)
            n        = len(coords) - 1

            if n < 3:
                continue

            # ── Non-convex spike ──────────────────────────────────
            if check_nonconvex:
                hull_geom  = feature.geometry().convexHull()
                hull_area  = hull_geom.area() if hull_geom else 0
                feat_area  = feature.geometry().area()

                if hull_area > 0 and feat_area > 5:
                    hull_ratio = feat_area / hull_area
                    if hull_ratio < 0.4:
                        # Find the vertex furthest OUTSIDE the convex hull
                        # That is the actual spike tip — not the centroid
                        hull_wkt   = hull_geom.asWkt()
                        best_pt    = None
                        best_dist  = -1
                        best_vidx  = 0

                        for vi, coord in enumerate(coords):
                            pt_geom = QgsGeometry.fromPointXY(
                                QgsPointXY(coord[0], coord[1])
                            )
                            # distance > 0 means outside hull
                            dist = pt_geom.distance(hull_geom)
                            if dist > best_dist:
                                best_dist = dist
                                best_pt   = coord
                                best_vidx = vi

                        # If no vertex outside hull (dist=0 for all),
                        # fall back to vertex with min angle
                        if best_pt is None or best_dist == 0:
                            # Find sharpest interior angle as fallback
                            min_angle = 360
                            best_pt   = coords[0]
                            for vi in range(n):
                                p1 = coords[(vi - 1) % n]
                                p2 = coords[vi]
                                p3 = coords[(vi + 1) % n]
                                v1 = (p1[0]-p2[0], p1[1]-p2[1])
                                v2 = (p3[0]-p2[0], p3[1]-p2[1])
                                l1 = math.sqrt(v1[0]**2 + v1[1]**2)
                                l2 = math.sqrt(v2[0]**2 + v2[1]**2)
                                if l1 > 0 and l2 > 0:
                                    cos_a = max(-1.0, min(1.0,
                                        (v1[0]*v2[0]+v1[1]*v2[1])/(l1*l2)))
                                    ang = math.degrees(math.acos(cos_a))
                                    if ang < min_angle:
                                        min_angle = ang
                                        best_pt   = coords[vi]
                                        best_vidx = vi

                        results.append({
                            "fid"        : fid,
                            "sl.no"      : slno,
                            "AiDashCode" : code,
                            "part"       : part_idx,
                            "vertex"     : best_vidx,
                            "angle"      : None,
                            "e_in"       : None,
                            "e_out"      : None,
                            "hull_ratio" : round(hull_ratio, 4),
                            "area_m2"    : round(feat_area, 2),
                            "x"          : best_pt[0],
                            "y"          : best_pt[1],
                            "error_cat"  : "NON-CONVEX SPIKE",
                            "error_type" : (
                                f"NON-CONVEX SPIKE "
                                f"hull_ratio={round(hull_ratio,3)} "
                                f"area={round(feat_area,2)}m² "
                                f"tip_vertex={best_vidx}"
                            ),
                        })

            # ── Sharp spike ───────────────────────────────────────
            if check_spikes:
                for i in range(n):
                    p1 = coords[(i - 1) % n]
                    p2 = coords[i]
                    p3 = coords[(i + 1) % n]

                    v1 = (p1[0]-p2[0], p1[1]-p2[1])
                    v2 = (p3[0]-p2[0], p3[1]-p2[1])
                    l1 = math.sqrt(v1[0]**2 + v1[1]**2)
                    l2 = math.sqrt(v2[0]**2 + v2[1]**2)

                    if l1 > 0 and l2 > 0:
                        cos_a = max(-1.0, min(1.0,
                                    (v1[0]*v2[0]+v1[1]*v2[1]) / (l1*l2)))
                        angle = math.degrees(math.acos(cos_a))

                        if angle < angle_threshold:
                            results.append({
                                "fid"        : fid,
                                "sl.no"      : slno,
                                "AiDashCode" : code,
                                "part"       : part_idx,
                                "vertex"     : i,
                                "angle"      : round(angle, 4),
                                "e_in"       : round(l1, 3),
                                "e_out"      : round(l2, 3),
                                "hull_ratio" : None,
                                "area_m2"    : None,
                                "x"          : p2[0],
                                "y"          : p2[1],
                                "error_cat"  : "SHARP SPIKE",
                                "error_type" : (
                                    f"SHARP SPIKE v{i} "
                                    f"angle={round(angle,4)}° "
                                    f"e_in={round(l1,3)}m "
                                    f"e_out={round(l2,3)}m"
                                ),
                            })

    return results


# ── Worker thread ─────────────────────────────────────────────────

class DetectionWorker(QThread):
    finished = pyqtSignal(list)
    progress = pyqtSignal(int, str)
    error    = pyqtSignal(str)

    def __init__(self, layer, threshold, spikes, nonconvex):
        super().__init__()
        self.layer     = layer
        self.threshold = threshold
        self.spikes    = spikes
        self.nonconvex = nonconvex

    def run(self):
        try:
            self.progress.emit(10, "Analysing features...")
            results = find_errors_in_layer(
                self.layer, self.threshold,
                self.spikes, self.nonconvex
            )
            self.progress.emit(100, f"Done — {len(results)} errors found")
            self.finished.emit(results)
        except Exception as e:
            self.error.emit(str(e))


# ── Main dialog ───────────────────────────────────────────────────

class SpikeCheckerDialog(QDialog):

    def __init__(self, iface):
        super().__init__(iface.mainWindow())
        self.iface       = iface
        self.layer       = None
        self.results     = []
        self.point_layer = None
        self.worker      = None

        self.setWindowTitle("Spike & Non-Convex Error Checker")
        self.setMinimumSize(960, 640)
        self._build_ui()
        self._apply_style()

    def _build_ui(self):
        main = QVBoxLayout(self)
        main.setSpacing(8)
        main.setContentsMargins(12, 12, 12, 12)

        hdr = QLabel("🔍  Spike & Non-Convex Error Checker")
        hdr.setObjectName("header")
        main.addWidget(hdr)

        # Input group
        ig = QGroupBox("Input")
        il = QVBoxLayout(ig)

        file_row = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        self.file_label.setObjectName("fileLabel")
        self.file_label.setWordWrap(True)
        browse_btn = QPushButton("📂  Browse Shapefile")
        browse_btn.setObjectName("browseBtn")
        browse_btn.setFixedWidth(190)
        browse_btn.clicked.connect(self._browse_file)
        file_row.addWidget(self.file_label, 1)
        file_row.addWidget(browse_btn)
        il.addLayout(file_row)

        opts_row = QHBoxLayout()
        opts_row.addWidget(QLabel("Angle threshold (°):"))
        self.thresh_input = QLineEdit("2.0")
        self.thresh_input.setFixedWidth(70)
        self.thresh_input.setToolTip(
            "Flag spike vertices with angle below this value.\n"
            "Genuine spikes are typically < 1°."
        )
        opts_row.addWidget(self.thresh_input)
        opts_row.addSpacing(20)

        self.cb_spikes    = QCheckBox("Sharp Spikes (angle < threshold)")
        self.cb_spikes.setChecked(True)
        self.cb_nonconvex = QCheckBox("Non-Convex Spikes (hull ratio < 0.4)")
        self.cb_nonconvex.setChecked(True)
        opts_row.addWidget(self.cb_spikes)
        opts_row.addSpacing(10)
        opts_row.addWidget(self.cb_nonconvex)
        opts_row.addStretch()
        il.addLayout(opts_row)
        main.addWidget(ig)

        # Buttons
        btn_row = QHBoxLayout()
        self.run_btn = QPushButton("▶  Run Detection")
        self.run_btn.setObjectName("runBtn")
        self.run_btn.setEnabled(False)
        self.run_btn.setFixedHeight(38)
        self.run_btn.clicked.connect(self._run_detection)

        self.load_btn = QPushButton("📍  Load Points to QGIS")
        self.load_btn.setObjectName("loadBtn")
        self.load_btn.setEnabled(False)
        self.load_btn.setFixedHeight(38)
        self.load_btn.clicked.connect(self._load_points)

        self.export_btn = QPushButton("💾  Export Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setEnabled(False)
        self.export_btn.setFixedHeight(38)
        self.export_btn.clicked.connect(self._export_excel)

        btn_row.addWidget(self.run_btn, 2)
        btn_row.addWidget(self.load_btn, 1)
        btn_row.addWidget(self.export_btn, 1)
        main.addLayout(btn_row)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setFixedHeight(6)
        self.progress_bar.setTextVisible(False)
        main.addWidget(self.progress_bar)

        self.summary_label = QLabel("")
        self.summary_label.setObjectName("summaryLabel")
        main.addWidget(self.summary_label)

        rg = QGroupBox("Errors Found")
        rl = QVBoxLayout(rg)
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "sl.no", "AiDashCode", "Error Type",
            "Vertex", "Angle (°)", "Edge In (m)", "Edge Out (m)",
            "Coordinates"
        ])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.itemSelectionChanged.connect(self._on_row_select)
        rl.addWidget(self.table)
        main.addWidget(rg, 1)

        self.status = QLabel("Ready — browse a polygon shapefile to begin.")
        self.status.setObjectName("statusBar")
        main.addWidget(self.status)

    def _apply_style(self):
        self.setStyleSheet("""
            QDialog {
                background-color: #1e1e2e; color: #cdd6f4;
                font-family: 'Segoe UI', sans-serif; font-size: 12px;
            }
            QLabel#header {
                font-size: 17px; font-weight: bold;
                color: #89b4fa; padding: 4px 0 8px 0;
            }
            QGroupBox {
                border: 1px solid #313244; border-radius: 6px;
                margin-top: 10px; padding-top: 10px;
                color: #a6adc8; font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 10px;
                padding: 0 4px; color: #89b4fa;
            }
            QLabel { color: #cdd6f4; }
            QLabel#fileLabel {
                color: #a6adc8; font-style: italic; padding: 4px 8px;
                background: #181825; border: 1px solid #313244;
                border-radius: 4px; min-height: 24px;
            }
            QLabel#summaryLabel {
                color: #a6e3a1; font-weight: bold; padding: 2px 4px;
            }
            QLabel#statusBar {
                color: #6c7086; font-size: 11px; padding: 2px 4px;
                border-top: 1px solid #313244;
            }
            QCheckBox { color: #cdd6f4; spacing: 6px; }
            QCheckBox::indicator {
                width: 14px; height: 14px;
                border: 1px solid #45475a; border-radius: 3px;
                background: #181825;
            }
            QCheckBox::indicator:checked { background: #89b4fa; border-color: #89b4fa; }
            QPushButton {
                background: #313244; color: #cdd6f4;
                border: 1px solid #45475a; border-radius: 5px; padding: 6px 14px;
            }
            QPushButton:hover { background: #45475a; border-color: #89b4fa; }
            QPushButton:disabled { color: #585b70; border-color: #313244; background: #1e1e2e; }
            QPushButton#browseBtn { background: #45475a; }
            QPushButton#runBtn {
                background: #89b4fa; color: #1e1e2e; font-weight: bold; border: none;
            }
            QPushButton#runBtn:hover { background: #b4befe; }
            QPushButton#runBtn:disabled { background: #313244; color: #585b70; }
            QPushButton#loadBtn {
                background: #a6e3a1; color: #1e1e2e; font-weight: bold; border: none;
            }
            QPushButton#loadBtn:hover { background: #94e2d5; }
            QPushButton#exportBtn {
                background: #f9e2af; color: #1e1e2e; font-weight: bold; border: none;
            }
            QLineEdit {
                background: #181825; color: #cdd6f4;
                border: 1px solid #45475a; border-radius: 4px; padding: 4px 8px;
            }
            QTableWidget {
                background: #181825; color: #cdd6f4;
                gridline-color: #313244; border: 1px solid #313244;
                border-radius: 4px; alternate-background-color: #1e1e2e;
                selection-background-color: #45475a;
            }
            QHeaderView::section {
                background: #313244; color: #89b4fa;
                border: none; padding: 6px 8px; font-weight: bold;
            }
            QProgressBar { background: #313244; border: none; border-radius: 3px; }
            QProgressBar::chunk { background: #89b4fa; border-radius: 3px; }
        """)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Polygon Shapefile", "", "Shapefiles (*.shp)"
        )
        if not path:
            return
        layer = QgsVectorLayer(path, os.path.basename(path), "ogr")
        if not layer.isValid():
            QMessageBox.critical(self, "Error", f"Could not load:\n{path}")
            return
        if QgsWkbTypes.geometryType(layer.wkbType()) != QgsWkbTypes.PolygonGeometry:
            QMessageBox.warning(self, "Wrong Type",
                                "Please select a POLYGON shapefile.")
            return
        self.layer = layer
        self.file_label.setText(path)
        self.run_btn.setEnabled(True)
        self.summary_label.setText("")
        self.table.setRowCount(0)
        self._set_status(
            f"Loaded: {layer.name()} — {layer.featureCount()} features"
        )

    def _run_detection(self):
        if self.layer is None:
            return
        try:
            threshold = float(self.thresh_input.text())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input",
                                "Angle threshold must be a number.")
            return
        if not self.cb_spikes.isChecked() and not self.cb_nonconvex.isChecked():
            QMessageBox.warning(self, "Nothing selected",
                                "Enable at least one check.")
            return

        self.run_btn.setEnabled(False)
        self.load_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.table.setRowCount(0)
        self.results = []
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(10)
        self._set_status("Running detection...")

        self.worker = DetectionWorker(
            self.layer, threshold,
            self.cb_spikes.isChecked(),
            self.cb_nonconvex.isChecked()
        )
        self.worker.finished.connect(self._on_finished)
        self.worker.progress.connect(
            lambda v, m: (self.progress_bar.setValue(v), self._set_status(m))
        )
        self.worker.error.connect(self._on_error)
        self.worker.start()

    def _on_finished(self, results):
        self.results = results
        self.progress_bar.setValue(100)
        self.progress_bar.setVisible(False)
        self.run_btn.setEnabled(True)

        if not results:
            self.summary_label.setText("✅  No errors found!")
            self._set_status("Detection complete — no errors.")
            return

        self.table.setRowCount(len(results))
        for row, r in enumerate(results):
            self.table.setItem(row, 0, QTableWidgetItem(str(r["sl.no"])))
            self.table.setItem(row, 1, QTableWidgetItem(str(r["AiDashCode"])))

            cat_item = QTableWidgetItem(r["error_cat"])
            cat_item.setForeground(QColor(
                "#f38ba8" if r["error_cat"] == "SHARP SPIKE" else "#fab387"
            ))
            self.table.setItem(row, 2, cat_item)
            self.table.setItem(row, 3, QTableWidgetItem(str(r["vertex"])))

            angle_str  = str(r["angle"]) if r["angle"] is not None else "—"
            angle_item = QTableWidgetItem(angle_str)
            if r["angle"] is not None:
                angle_item.setForeground(QColor(
                    "#f38ba8" if r["angle"] < 0.1 else
                    "#fab387" if r["angle"] < 1.0 else "#f9e2af"
                ))
            self.table.setItem(row, 4, angle_item)
            self.table.setItem(row, 5, QTableWidgetItem(
                str(r["e_in"])  if r["e_in"]  is not None else "—"))
            self.table.setItem(row, 6, QTableWidgetItem(
                str(r["e_out"]) if r["e_out"] is not None else "—"))
            self.table.setItem(row, 7, QTableWidgetItem(
                f"({round(r['x'],2)}, {round(r['y'],2)})"))

        spikes_n    = sum(1 for r in results if r["error_cat"] == "SHARP SPIKE")
        nonconvex_n = sum(1 for r in results if r["error_cat"] == "NON-CONVEX SPIKE")
        unique_poly = len(set(r["sl.no"] for r in results))
        parts = []
        if spikes_n:    parts.append(f"{spikes_n} sharp spike(s)")
        if nonconvex_n: parts.append(f"{nonconvex_n} non-convex spike(s)")
        self.summary_label.setText(
            f"⚠  {' + '.join(parts)} across {unique_poly} polygon(s)"
        )
        self.load_btn.setEnabled(True)
        self.export_btn.setEnabled(True)
        self._set_status(
            f"Done — {len(results)} errors. "
            "Click a row to zoom, or load all points."
        )

    def _on_error(self, msg):
        self.progress_bar.setVisible(False)
        self.run_btn.setEnabled(True)
        QMessageBox.critical(self, "Error", f"Detection failed:\n{msg}")
        self._set_status(f"Error: {msg}")

    def _on_row_select(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self.results):
            return
        r = self.results[row]
        canvas = self.iface.mapCanvas()
        canvas.setCenter(QgsPointXY(r["x"], r["y"]))
        canvas.zoomScale(500)
        canvas.refresh()

    def _load_points(self):
        """Load in-memory point layer — NO shp saved to disk."""
        if not self.results:
            return

        # Remove previous layer if exists
        if self.point_layer is not None:
            try:
                QgsProject.instance().removeMapLayer(self.point_layer.id())
            except Exception:
                pass
            self.point_layer = None

        crs_str   = self.layer.crs().authid() or "EPSG:27700"
        mem_layer = QgsVectorLayer(
            f"Point?crs={crs_str}", "Spike_Errors", "memory"
        )
        provider = mem_layer.dataProvider()
        provider.addAttributes([
            QgsField("sl_no",      QVariant.String),
            QgsField("AiDashCode", QVariant.String),
            QgsField("error_cat",  QVariant.String),
            QgsField("vertex",     QVariant.String),
            QgsField("angle_deg",  QVariant.Double),
            QgsField("e_in_m",     QVariant.Double),
            QgsField("e_out_m",    QVariant.Double),
            QgsField("error_type", QVariant.String),
        ])
        mem_layer.updateFields()

        features = []
        for r in self.results:
            f = QgsFeature()
            f.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(r["x"], r["y"])))
            f.setAttributes([
                str(r["sl.no"]),
                str(r["AiDashCode"]),
                r["error_cat"],
                str(r["vertex"]),
                r["angle"]  if r["angle"] is not None else None,
                r["e_in"]   if r["e_in"]  is not None else None,
                r["e_out"]  if r["e_out"] is not None else None,
                r["error_type"],
            ])
            features.append(f)

        provider.addFeatures(features)
        mem_layer.updateExtents()

        # ── Simple red circle, 1mm ────────────────────────────────
        symbol = QgsMarkerSymbol.createSimple({
            "name"              : "circle",
            "color"             : "#e74c3c",
            "color_border"      : "#c0392b",
            "size"              : "1",
            "size_unit"         : "MM",
            "outline_width"     : "0.2",
            "outline_width_unit": "MM",
        })
        mem_layer.setRenderer(QgsSingleSymbolRenderer(symbol))
        mem_layer.triggerRepaint()

        QgsProject.instance().addMapLayer(mem_layer)
        self.point_layer = mem_layer
        self._set_status(
            f"✅  {len(self.results)} points loaded as 'Spike_Errors' "
            f"(in-memory — not saved to disk)"
        )

    def _export_excel(self):
        if not self.results:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel Report",
            os.path.dirname(self.file_label.text()),
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Errors"

            headers = [
                "sl.no", "AiDashCode", "Error Category", "Vertex",
                "Angle (°)", "Edge In (m)", "Edge Out (m)",
                "Hull Ratio", "Area (m²)", "X", "Y", "Error Detail"
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill      = PatternFill("solid", fgColor="1F4E79")
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

            for row, r in enumerate(self.results, 2):
                ws.cell(row=row, column=1,  value=str(r["sl.no"]))
                ws.cell(row=row, column=2,  value=str(r["AiDashCode"]))
                ws.cell(row=row, column=3,  value=r["error_cat"])
                ws.cell(row=row, column=4,  value=str(r["vertex"]))
                ws.cell(row=row, column=5,  value=r["angle"])
                ws.cell(row=row, column=6,  value=r["e_in"])
                ws.cell(row=row, column=7,  value=r["e_out"])
                ws.cell(row=row, column=8,  value=r["hull_ratio"])
                ws.cell(row=row, column=9,  value=r["area_m2"])
                ws.cell(row=row, column=10, value=round(r["x"], 4))
                ws.cell(row=row, column=11, value=round(r["y"], 4))
                ws.cell(row=row, column=12, value=r["error_type"])

                fill = (
                    PatternFill("solid", fgColor="FFC7CE")
                    if r["error_cat"] == "SHARP SPIKE"
                    else PatternFill("solid", fgColor="FFEB9C")
                )
                for col in range(1, 13):
                    ws.cell(row=row, column=col).fill = fill

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            wb.save(path)
            self._set_status(f"Excel saved → {os.path.basename(path)}")
            QMessageBox.information(self, "Saved", f"Saved:\n{path}")
        except ImportError:
            QMessageBox.warning(self, "Missing Library",
                                "Install openpyxl: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not save:\n{e}")

    def _set_status(self, msg):
        self.status.setText(msg)
